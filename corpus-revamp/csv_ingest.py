"""csv_ingest.py — corpus-builder entry point.

The user hands a CSV or Excel of videos (IDs, full URLs, or titles). This:
  1. parses every cell, extracting YouTube video IDs (from IDs or URLs);
  2. if a row has no ID but has a title, optionally resolves it via YouTube search;
  3. dedupes against the existing corpus (shared_db + shared_db_v2) and within the list;
  4. runs each NEW video through the SAME v3 pipeline (reclassify.ingest_list);
  5. exports the updated clean corpus to CSV.

Resumable + crash-safe (a finished video is skipped on re-run). Claude does NOT source
videos itself — the user provides them — unless --titles is given to search.

Usage:
  python csv_ingest.py videos.csv                 # use IDs/URLs found in the file
  python csv_ingest.py videos.xlsx --titles 2     # also search 2 videos per title cell
  python csv_ingest.py --selftest                 # preflight (detector + VLM + a download)
"""
from __future__ import annotations
import sys, os, re, json, glob, csv as csvmod

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
OUTPUTS = os.path.abspath(os.path.join(HERE, "..", "outputs"))

_VID = re.compile(r"(?:v=|youtu\.be/|/watch\?v=|/shorts/|/embed/)([A-Za-z0-9_-]{11})")
_BARE = re.compile(r"^[A-Za-z0-9_-]{11}$")


def _read_rows(path):
    if path.lower().endswith((".xlsx", ".xlsm", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            for r in ws.iter_rows(values_only=True):
                yield ["" if c is None else str(c) for c in r]
    else:
        with open(path, newline="", encoding="utf-8-sig") as fh:
            for r in csvmod.reader(fh):
                yield r


def _dedupe(xs):
    seen = set()
    return [x for x in xs if not (x in seen or seen.add(x))]


def parse_file(path):
    """Return (ids, titles) found in the file (order-preserving, deduped)."""
    ids, titles = [], []
    for r in _read_rows(path):
        row_has_id = False
        for cell in r:
            s = (cell or "").strip()
            if not s:
                continue
            m = _VID.search(s)
            if m:
                ids.append(m.group(1)); row_has_id = True
            elif _BARE.match(s):
                ids.append(s); row_has_id = True
        if not row_has_id:
            for cell in r:
                t = (cell or "").strip()
                if len(t) >= 12 and " " in t and any(c.isalpha() for c in t) and "http" not in t.lower():
                    titles.append(t); break
    return _dedupe(ids), _dedupe(titles)


def existing_ids():
    ids = set()
    for p in (glob.glob(OUTPUTS + "/shared_db/records/*.json")
              + glob.glob(OUTPUTS + "/shared_db_v2/records/*.json")
              + glob.glob(OUTPUTS + "/shared_db_v2/rejected/*.json")):
        try:
            ids.add(json.load(open(p)).get("video_id"))
        except Exception:
            pass
    return ids


def main():
    if "--selftest" in sys.argv:
        import reclassify
        return reclassify.selftest()
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if not args:
        print(__doc__); sys.exit(1)
    path = args[0]
    per_title = 0
    if "--titles" in sys.argv:
        i = sys.argv.index("--titles")
        per_title = int(sys.argv[i + 1]) if i + 1 < len(sys.argv) and sys.argv[i + 1].isdigit() else 2

    ids, titles = parse_file(path)
    print(f"parsed {path}: {len(ids)} video ids, {len(titles)} title rows")
    have = existing_ids()

    # resolve titles -> ids via search (only if --titles)
    if per_title and titles:
        import search_titles  # reuses its keyword + ytsearch + dedupe
        seen = set(have) | set(ids)
        for t in titles:
            try:
                found = search_titles.search(search_titles.keywords(t), per_title)
            except Exception as e:
                print(f"  [searchfail] {t[:40]}: {str(e)[:50]}"); continue
            for v in found:
                if v not in seen:
                    seen.add(v); ids.append(v)
        print(f"after title search (+{per_title}/title): {len(ids)} candidate ids")

    todo = [v for v in _dedupe(ids) if v not in have]
    print(f"NEW videos to ingest (deduped vs corpus): {len(todo)} (skipping {len(ids) - len(todo)} already present)")
    if not todo:
        print("nothing new to ingest.")
    else:
        json.dump(todo, open("/tmp/csv_ingest_ids.json", "w"))
        import reclassify
        reclassify.ingest_list(todo, source="csv-ingest")

    # always (re)export the full clean corpus CSV
    import export_csv
    out = os.path.join(OUTPUTS, "shared_db_v2", "corpus_clean_windows.csv")
    n = 0
    with open(out, "w", newline="", encoding="utf-8") as fh:
        wr = csvmod.DictWriter(fh, fieldnames=export_csv.FIELDS)
        wr.writeheader()
        for row in export_csv.iter_rows():
            wr.writerow(row); n += 1
    print(f"corpus now {n} clean windows -> {out}")


if __name__ == "__main__":
    main()
